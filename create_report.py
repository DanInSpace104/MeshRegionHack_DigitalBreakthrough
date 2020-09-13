import django
import datetime
from openpyxl import Workbook
import os
import xml.etree.ElementTree as ET


def create_report(date):
    '''
    Функция создания отчета, для дальнейшего вывода или отправки
    format - xls, xml, xlsx
    '''
    ID = [i['id'] for i in date]
    company_name = [i['company_name'] for i in date]
    INN = [i['inn'] for i in date]
    KPP = [i['kpp'] for i in date]
    go_fl = [i['gofl'] for i in date]

    bik = [i['bik'] for i in date]
    bank_name = [i['bank_name'] for i in date]

    comment = [i['comment'] for i in date]
    acc_type = [i['acc_type'] for i in date]
    contract_create_date = [i['contract_create_date'] for i in date]
    contract_begin_date = [i['contract_begin_date'] for i in date]
    contract_end_date = [i['contract_end_date'] for i in date]
    settlement_rate = [i['settlement_rate'] for i in date]
    currency_code = [i['currency_code'] for i in date]
    summ = [i['summ'] for i in date]
    balance = [i['balance'] for i in date]
    
    now = datetime.datetime.now()  # Текущая дата
    # now_format = now.strftime("%Y-%m-%d_%H-%M")
    now_format = now.strftime("%Y-%m-%d")
    directory = os.path.abspath(os.curdir)  # Директория
    
    # Создаем подключение
    wb = Workbook()
    ws = wb.active

    
    ws.merge_cells('B4:D4')
    ws.merge_cells('B5:D5')
    ws['A4'] = 'Организация'
    ws['A5'] = 'ИНН'
    ws['B4'] = str(company_name[0])
    ws['B5'] = str(INN[0])

    ws.merge_cells('B7:K7')

    ws.merge_cells('A7:A10')
    ws.merge_cells('B8:B10')
    ws.merge_cells('C8:C10')
    ws.merge_cells('D8:D10')
    ws.merge_cells('E8:E10')
    ws.merge_cells('F8:F10')
    ws.merge_cells('G8:G10')
    ws.merge_cells('H8:H10')
    ws.merge_cells('I8:I10')
    ws.merge_cells('J8:J9')
    ws.merge_cells('K8:K9')
    ws.merge_cells('M9:M10')
    ws.merge_cells('N9:N10')
    ws.merge_cells('O9:O10')
    ws.merge_cells('P9:P10')
    ws.merge_cells('R9:R10')
    ws.merge_cells('L8:R8')
    ws.merge_cells('L7:R7')
    ws['A7'] = '№'
    ws['B7'] = 'К О Н Т Р А Г Е Н Т Ы  / О С Т А Т К И     Н А    С Ч Е Т А Х'
    ws['L7'] = 'РАЗМЕЩЕНИЕ И ПРИВЛЕЧЕНИЕ СРЕДСТВ'
    ws['B8'] = 'ID'
    ws['C8'] = 'ИНН'
    ws['D8'] = 'КПП'
    ws['E8'] = 'Наименование организации'
    ws['F8'] = 'ГО/ФЛ'
    ws['G8'] = 'БИК банка'
    ws['H8'] = 'Наименование банка'
    ws['I8'] = 'Комментарий'
    ws['J8'] = 'Валюта'
    ws['J10'] = 'международный код'
    ws['K8'] = 'Сумма в валюте счета'
    ws['K10'] = 'Остаток на счете'
    ws['L9'] = 'ВИД договора'
    ws['M9'] = 'Даиа заключение договора'
    ws['N9'] = 'Дата начала действия договора'
    ws['O9'] = 'Дата окончания договора'
    ws['P9'] = 'Расчитанная ставка в годовых, %'
    ws['Q9'] = 'Валюта'
    ws['Q10'] = 'Международный код'
    ws['R10'] = 'Сумма'
    # ws.append([1, 2, 3])
    for i in range(len(date)):
        ws[f'A{12+i}'] = f'{i}'
        ws[f'B{12+i}'] = f'{ID[i]}'
        ws[f'C{12+i}'] = f'{INN[i]}'
        ws[f'D{12+i}'] = f'{KPP[i]}'
        ws[f'E{12+i}'] = f'{company_name[i]}'
        ws[f'F{12+i}'] = f'{go_fl[i]}'
        ws[f'G{12+i}'] = f'{bik[i]}'
        ws[f'H{12+i}'] = f'{bank_name[i]}'
        ws[f'I{12+i}'] = f'{comment[i]}'
        ws[f'J{12+i}'] = f'{currency_code[i]}'
        ws[f'K{12+i}'] = f'{balance[i]}'
        ws[f'L{12+i}'] = f'{acc_type[i]}'
        ws[f'M{12+i}'] = f'{contract_create_date[i]}'
        ws[f'N{12+i}'] = f'{contract_begin_date[i]}'
        ws[f'O{12+i}'] = f'{contract_end_date[i]}'
        ws[f'P{12+i}'] = f'{settlement_rate[i]}'
        ws[f'Q{12+i}'] = f'{currency_code[i]}'
        ws[f'R{12+i}'] = f'{summ[i]}'

    global_class = ET.Element('Request')
    for i in range(len(date)):
        data = ET.SubElement(global_class, f'Report_{i+1}')

        # Создаются блоки
        organ_block = ET.SubElement(data, 'Organization')
        bank_block = ET.SubElement(data, 'Bank')
        score_block = ET.SubElement(data, 'Score')

        # Создаем элементы
        ID_el = ET.SubElement(organ_block, f'ID')
        INN_el = ET.SubElement(organ_block, 'INN')
        KPP_el = ET.SubElement(organ_block, 'KPP')
        go_fl_el = ET.SubElement(organ_block, 'go-fl')

        bik_el = ET.SubElement(bank_block, 'bik')
        bank_name_el = ET.SubElement(bank_block, 'bank_name')

        acc_type_el = ET.SubElement(score_block, 'acc_type')
        contract_create_date_el = ET.SubElement(score_block, 'contract_create_date')
        contract_begin_date_el = ET.SubElement(score_block, 'contract_begin_date')
        contract_end_date_el = ET.SubElement(score_block, 'contract_end_date')
        settlement_rate_el = ET.SubElement(score_block, 'settlement_rate')
        currency_code_el = ET.SubElement(score_block, 'currency_code')
        summ_el = ET.SubElement(score_block, 'summ')
        balance_el = ET.SubElement(score_block, 'balance')
        comment_el = ET.SubElement(score_block, 'comment')

        # Наполняем элементы
        ID_el.text = str(ID[i])
        INN_el.text = str(INN[i])
        KPP_el.text = str(KPP[i])
        go_fl_el.text = str(go_fl[i])

        bik_el.text = str(bik[i])
        bank_name_el.text = str(bank_name[i])

        acc_type_el.text = str(acc_type[i])
        contract_create_date_el.text = str(contract_create_date[i])
        contract_begin_date_el.text = str(contract_begin_date[i])
        contract_end_date_el.text = str(contract_end_date[i])
        settlement_rate_el.text = str(settlement_rate[i])
        currency_code_el.text = str(currency_code[i])
        summ_el.text = str(summ[i])
        balance_el.text = str(balance[i])
        comment_el.text = str(comment[i])

    b_xml = ET.tostring(global_class)
    directory = os.path.abspath(os.curdir)
    save_name = str(company_name[0])+ '_' + str(now_format)
    save_dir = [directory+save_name+'.xlsx',
                directory+save_name+'.xls',
                directory+save_name+'.xml']
    with open(f"{save_name}.xml", "wb") as f:
        f.write(b_xml)
    wb.save(f"{save_name}.xls")
    wb.save(f"{save_name}.xlsx")
    return (save_dir)


date = [
    {
        "id": 1,
        "currency_code": "RUB",
        "bank_name": "ВТБ",
        "bik": 123412341234,
        "kpp": 123123,
        "inn": 1231231,
        "company_name": "Jandex",
        "gofl": "ГО",
        "acc_type": "Депозит",
        "balance": 1666.0,
        "summ": 200000.0,
        "contract_create_date": "2020-09-13",
        "contract_begin_date": "2020-09-13",
        "comment": "",
        "contract_end_date": "2020-09-13",
        "settlement_rate": 15.0,
        "bank": 1,
        "company": 1,
        "currency": 1
    },
    {
        "id": 2,
        "currency_code": "RUB",
        "bank_name": "ВТБ",
        "bik": 123412341234,
        "kpp": 123123,
        "inn": 1231231,
        "company_name": "Jandex",
        "gofl": "ГО",
        "acc_type": "Депозит",
        "balance": 1666.0,
        "summ": 200000.0,
        "contract_create_date": "2020-09-13",
        "contract_begin_date": "2020-09-13",
        "comment": "",
        "contract_end_date": "2020-09-13",
        "settlement_rate": 15.0,
        "bank": 1,
        "company": 1,
        "currency": 1
    },
    {
        "id": 3,
        "currency_code": "USD",
        "bank_name": "Сбербанк",
        "bik": 9191919,
        "kpp": 123123,
        "inn": 1231231,
        "company_name": "Jandex",
        "gofl": "ГО",
        "acc_type": "Кредит",
        "balance": 999.0,
        "summ": 888.0,
        "contract_create_date": "2020-09-13",
        "contract_begin_date": "2020-09-13",
        "comment": "",
        "contract_end_date": "2020-09-13",
        "settlement_rate": 23.0,
        "bank": 2,
        "company": 1,
        "currency": 2
    }
]

create_report(date)
